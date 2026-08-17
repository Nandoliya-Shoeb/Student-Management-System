from django.http import HttpResponse
from django.utils.translation import gettext as _
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Q, Count, Avg
from .models import Attendance, Fee, Student


def generate_pdf_receipt(fee):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receipt_{fee.receipt_number}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=30,
        alignment=1,
    )
    
    story.append(Paragraph(_('Fee Receipt'), title_style))
    story.append(Spacer(1, 0.3*inch))
    
    data = [
        [_('Receipt Number:'), fee.receipt_number],
        [_('Student Name:'), fee.student.name],
        [_('Student ID:'), fee.student.student_id],
        [_('Amount:'), f"₹ {fee.amount}"],
        [_('Fee Type:'), fee.get_fee_type_display()],
        [_('Status:'), fee.get_status_display()],
        [_('Due Date:'), str(fee.due_date)],
        [_('Payment Date:'), str(fee.payment_date) if fee.payment_date else _('Pending')],
        [_('Payment Method:'), fee.get_payment_method_display() if fee.payment_method else _('N/A')],
    ]
    
    table = Table(data, colWidths=[2*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(table)
    
    doc.build(story)
    return response


def generate_pdf_attendance(attendance_records, start_date, end_date=None):
    if end_date is None:
        end_date = start_date

    is_range = (start_date != end_date)
    response = HttpResponse(content_type='application/pdf')
    if is_range:
        filename = f"attendance_{start_date}_to_{end_date}.pdf"
        title_text = f"{_('Attendance Report')} : {start_date.strftime('%d-%m-%Y')} {_('to')} {end_date.strftime('%d-%m-%Y')}"
    else:
        filename = f"attendance_{start_date}.pdf"
        title_text = f"{_('Attendance Report')} - {start_date.strftime('%d-%m-%Y')}"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=15,
        alignment=1,
    )

    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 0.2*inch))

    if is_range:
        data = [[_('Date'), _('Student ID'), _('Student Name'), _('Class'), _('Status'), _('Remarks')]]
        for record in attendance_records:
            data.append([
                record.date.strftime('%d-%m-%Y'),
                record.student.student_id,
                record.student.name,
                f"Grade {record.student.class_field}",
                record.get_status_display(),
                record.remarks or '',
            ])
        table = Table(data, colWidths=[1.1*inch, 1.1*inch, 1.6*inch, 0.8*inch, 0.9*inch, 1.3*inch])
    else:
        data = [[_('Student ID'), _('Student Name'), _('Class'), _('Status'), _('Remarks')]]
        for record in attendance_records:
            data.append([
                record.student.student_id,
                record.student.name,
                f"Grade {record.student.class_field}",
                record.get_status_display(),
                record.remarks or '',
            ])
        table = Table(data, colWidths=[1.1*inch, 1.8*inch, 0.9*inch, 1*inch, 1.6*inch])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    story.append(table)
    doc.build(story)
    return response


def generate_excel_report(report_type, data, date, response, end_date=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = report_type.capitalize()

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    is_range = end_date and (date != end_date)

    if report_type == 'attendance':
        if is_range:
            headers = [_('Date'), _('Student ID'), _('Student Name'), _('Class'), _('Status'), _('Remarks')]
            worksheet.append(headers)
            for record in data:
                worksheet.append([
                    record.date.strftime('%d-%m-%Y'),
                    record.student.student_id,
                    record.student.name,
                    f"Grade {record.student.class_field}",
                    str(record.get_status_display()),
                    record.remarks or '',
                ])
        else:
            headers = [_('Student ID'), _('Student Name'), _('Class'), _('Status'), _('Remarks')]
            worksheet.append(headers)
            for record in data:
                worksheet.append([
                    record.student.student_id,
                    record.student.name,
                    f"Grade {record.student.class_field}",
                    str(record.get_status_display()),
                    record.remarks or '',
                ])
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response


def calculate_attendance_percentage(student):
    total_days = Attendance.objects.filter(student=student).count()
    if total_days == 0:
        return 0
    
    present_days = Attendance.objects.filter(student=student, status='present').count()
    return round((present_days / total_days) * 100, 2)


def calculate_student_performance(student):
    from django.db.models import Avg, Max

    quiz_results = student.quiz_results.all()
    if not quiz_results.exists():
        return {
            'avg_score': 0,
            'best_score': 0,
            'total_attempts': 0,
            'passed_count': 0,
            'failed_count': 0,
        }

    avg_score = quiz_results.aggregate(avg=Avg('percentage'))['avg'] or 0
    # Bug 9 fix: was Avg('percentage') — must be Max
    best_score = quiz_results.aggregate(best=Max('percentage'))['best'] or 0
    passed_count = quiz_results.filter(passed=True).count()
    failed_count = quiz_results.filter(passed=False).count()

    return {
        'avg_score': round(float(avg_score), 2),
        'best_score': round(float(best_score), 2),
        'total_attempts': quiz_results.count(),
        'passed_count': passed_count,
        'failed_count': failed_count,
    }


def get_formatted_time(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    
    if hours > 0:
        return f"{hours}h {minutes}m {secs}s"
    elif minutes > 0:
        return f"{minutes}m {secs}s"
    else:
        return f"{secs}s"


def generate_pdf_result(quiz_result):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="result_{quiz_result.id}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=20,
        alignment=1,
    )
    
    story.append(Paragraph(_('Quiz Result'), title_style))
    story.append(Spacer(1, 0.3*inch))
    
    summary_data = [
        [_('Quiz Title:'), quiz_result.quiz.title],
        [_('Student Name:'), quiz_result.student.name],
        [_('Total Marks:'), f"{quiz_result.total_marks}"],
        [_('Percentage:'), f"{quiz_result.percentage}%"],
        [_('Status:'), _('Passed') if quiz_result.passed else _('Failed')],
        [_('Correct Answers:'), str(quiz_result.correct_answers)],
        [_('Wrong Answers:'), str(quiz_result.wrong_answers)],
        [_('Unanswered:'), str(quiz_result.unanswered)],
        [_('Time Taken:'), get_formatted_time(quiz_result.time_taken)],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(_('Question Details'), styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    details_data = [[_('Question'), _('Your Answer'), _('Correct Answer'), _('Result')]]
    
    for answer in quiz_result.student_answers.all():
        question = answer.question
        details_data.append([
            question.question_text[:50] + '...',
            answer.get_selected_option_display(),
            question.get_correct_option_display(),
            _('Correct') if answer.is_correct else _('Wrong') if answer.selected_option != 'unanswered' else _('Unanswered'),
        ])
    
    details_table = Table(details_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    story.append(details_table)
    
    doc.build(story)
    return response
